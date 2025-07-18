import os
import time
import openpyxl
import threading
from tkinter import Tk, Label, Button, filedialog, ttk, messagebox
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

class PSMEScraperApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PSME Member Scraper")
        self.root.geometry("400x200")

        self.label = Label(root, text="Select user Excel file to start.")
        self.label.pack(pady=10)

        self.select_button = Button(root, text="Choose Excel File", command=self.select_file)
        self.select_button.pack()

        self.start_button = Button(root, text="Start Scraping", command=self.start_scraping, state="disabled")
        self.start_button.pack(pady=10)

        self.progress = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
        self.progress.pack(pady=10)

        self.filepath = None

    def select_file(self):
        self.filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.filepath:
            self.label.config(text=f"Selected: {os.path.basename(self.filepath)}")
            self.start_button.config(state="normal")

    def start_scraping(self):
        threading.Thread(target=self.scrape_members).start()

    def scrape_members(self):
        try:
            user_wb = openpyxl.load_workbook(self.filepath)
            user_ws = user_wb.active
            users = list(user_ws.iter_rows(min_row=2, values_only=True))
            total = len(users)

            os.makedirs("output", exist_ok=True)
            summary_wb = openpyxl.Workbook()
            summary_ws = summary_wb.active
            summary_ws.title = "Summary"
            summary_ws.append(["user label", "username", "chapter name", "member count"])

            self.progress["maximum"] = total

            for i, (user_label, username, password) in enumerate(users, start=1):
                options = Options()
                options.add_argument("--headless")
                options.add_argument("--window-size=1920,1080")
                driver = webdriver.Chrome(options=options)

                try:
                    driver.get("https://psmeinc.org.ph/#/chapter-login")

                    WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.NAME, "username"))
                    ).send_keys(username)
                    driver.find_element(By.NAME, "password").send_keys(password)
                    driver.find_element(By.XPATH, "//button[contains(text(),'Login')]").click()

                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, "//span[contains(text(),'Chapter Management')]"))
                    ).click()
                    WebDriverWait(driver, 30).until(
                        EC.element_to_be_clickable((By.XPATH, "//a[contains(text(),'Chapter Members')]"))
                    ).click()

                    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//table")))
                    time.sleep(3)

                    rows = driver.find_elements(By.XPATH, "//table//tbody/tr")
                    member_count = len(rows)

                    try:
                        chapter_elem = driver.find_element(By.XPATH, "//div[contains(@class, 'text-sm') and contains(text(), 'Chapter')]")
                        chapter_name = chapter_elem.text.strip()
                    except:
                        chapter_name = "Unknown Chapter"

                    out_wb = openpyxl.Workbook()
                    out_ws = out_wb.active
                    out_ws.title = "Members"
                    out_ws.append(["chapter name", "member count"])
                    out_ws.append([chapter_name, member_count])

                    safe_name = user_label.replace(" ", "_").replace("/", "_")
                    out_wb.save(f"output/{safe_name}_count.xlsx")

                    summary_ws.append([user_label, username, chapter_name, member_count])

                except Exception as e:
                    print(f"Error for {username}: {e}")
                    summary_ws.append([user_label, username, "ERROR", 0])
                finally:
                    driver.quit()

                self.progress["value"] = i
                self.root.update_idletasks()

            summary_wb.save("output/master_summary.xlsx")
            messagebox.showinfo("Done", "Scraping completed!\nResults saved in the /output folder.")
            self.label.config(text="Process finished.")

        except Exception as e:
            messagebox.showerror("Error", str(e))


if __name__ == "__main__":
    root = Tk()
    app = PSMEScraperApp(root)
    root.mainloop()
